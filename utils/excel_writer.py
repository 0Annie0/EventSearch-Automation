# EventSearch/utils/excel_writer.py
import os
from typing import List, Tuple
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def write_results_to_excel(
    source_path: str,
    sheet_name: str,
    result_column: str,
    results: List[Tuple[int, str, bool]]
) -> str:
    """
    基于原 Excel 生成一份带测试结果的新文件。
    :param source_path: 原始 Excel 路径
    :param sheet_name: 工作表名称
    :param result_column: 结果列字母（如 'H'）
    :param results: 列表，元素为 (行号, 关键字, passed)
    :return: 生成的新文件路径
    """
    wb = load_workbook(source_path)
    ws = wb[sheet_name]

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    for row_num, keyword, passed in results:
        cell = ws[f"{result_column}{row_num}"]
        if passed:
            cell.value = "PASS"
        else:
            cell.value = "FAIL"
            cell.fill = red_fill

    base, ext = os.path.splitext(source_path)
    output_path = base + "_结果" + ext
    wb.save(output_path)
    print(f"测试结果已保存至: {output_path}")
    return output_path