# EventSearch/utils/excel_reader.py
from typing import List, Tuple
from openpyxl import load_workbook

def load_test_data(file_path: str, sheet_name: str, column_letter: str, start_row: int = 2) -> List[Tuple[int, str]]:
    """
    从 Excel 读取测试数据，返回 [(行号, 单元格值), ...]
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    data = []
    row = start_row
    while True:
        value = ws[f"{column_letter}{row}"].value
        if value is None or value == "":
            break
        keyword = str(value).strip()
        data.append((row, keyword))
        row += 1
    wb.close()
    return data